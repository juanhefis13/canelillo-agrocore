import json
import sys
from collections import defaultdict
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Falta openpyxl. Usa el Python bundled de Codex o instala openpyxl.") from exc


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def as_float(value, default=0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def as_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value:
        return str(value)[:10]
    return date.today().isoformat()


def slug(text, prefix):
    raw = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(text or prefix))
    raw = "-".join(part for part in raw.split("-") if part)
    return f"{prefix}-{raw[:48] or 'item'}"


BASE = {
    "settings": {
        "farmName": "Agricola El Canelillo",
        "season": "2024/2025",
        "defaultTankLiters": 2000,
        "lowStockDays": 21,
    },
    "products": [],
    "blocks": [],
    "operators": [],
    "equipment": [
        {"id": "eq-1", "type": "Nebulizadora", "code": "N-01", "tankLiters": 2000},
        {"id": "eq-2", "type": "Nebulizadora", "code": "N-02", "tankLiters": 1500},
        {"id": "tr-1", "type": "Tractor", "code": "T-01", "tankLiters": 0},
    ],
    "weather": [],
    "orders": [],
    "inventoryMovements": [],
}


def load_headers(ws, row_number):
    return {clean(cell.value): index for index, cell in enumerate(ws[row_number]) if clean(cell.value)}


def read_products(wb, state):
    if "Productos" not in wb.sheetnames:
        return {}
    ws = wb["Productos"]
    headers = load_headers(ws, 2)
    product_ids = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = clean(row[headers.get("NOMBRE COMERCIAL", 3)]) if "NOMBRE COMERCIAL" in headers else ""
        if not name or name in product_ids:
            continue
        product_id = slug(name, "p")
        product_ids[name.lower()] = product_id
        state["products"].append({
            "id": product_id,
            "name": name,
            "ingredient": clean(row[headers.get("INGREDIENTE ACTIVO", 4)]) if "INGREDIENTE ACTIVO" in headers else "",
            "unit": "L/kg",
            "dose100": as_float(row[headers.get("Dosis /100lts    (grs-cc)", 8)]) if "Dosis /100lts    (grs-cc)" in headers else 0,
            "reentryHours": 24,
            "carencyDays": as_float(row[headers.get("Período Carencia Agenda de Pesticidas", 7)]) if "Período Carencia Agenda de Pesticidas" in headers else 0,
            "stock": 0,
            "minStock": 0,
            "cost": 0,
            "lot": "",
            "expires": "",
        })
    return product_ids


def read_blocks(wb, state):
    if "HAS Canelillo" not in wb.sheetnames:
        return
    ws = wb["HAS Canelillo"]
    headers = load_headers(ws, 2)
    for row in ws.iter_rows(min_row=3, values_only=True):
        potrero = clean(row[headers.get("Potrero", 1)]) if "Potrero" in headers else ""
        block = clean(row[headers.get("Bloque", 2)]) if "Bloque" in headers else ""
        if not potrero or not block:
            continue
        state["blocks"].append({
            "id": slug(f"{potrero}-{block}", "b"),
            "potrero": potrero,
            "block": str(block),
            "crop": clean(row[headers.get("Especie", 3)]) if "Especie" in headers else "",
            "variety": clean(row[headers.get("Variedad", 4)]) if "Variedad" in headers else "",
            "hectares": as_float(row[headers.get("Hectareas", 5)]) if "Hectareas" in headers else 0,
        })


def ensure_product(state, product_ids, name, ingredient="", dose100=0):
    key = str(name).strip().lower()
    if not key:
        return ""
    if key in product_ids:
        return product_ids[key]
    product_id = slug(name, "p")
    product_ids[key] = product_id
    state["products"].append({
        "id": product_id,
        "name": str(name).strip(),
        "ingredient": str(ingredient or "").strip(),
        "unit": "L/kg",
        "dose100": as_float(dose100),
        "reentryHours": 24,
        "carencyDays": 0,
        "stock": 0,
        "minStock": 0,
        "cost": 0,
        "lot": "",
        "expires": "",
    })
    return product_id


def read_orders(wb, state, product_ids):
    if "Aplicaciones" not in wb.sheetnames:
        return
    ws = wb["Aplicaciones"]
    headers = load_headers(ws, 5)
    orders = {}
    operators = {}

    def value(row, header):
        index = headers.get(header)
        return row[index] if index is not None and index < len(row) else None

    for row in ws.iter_rows(min_row=6, values_only=True):
        number = clean(value(row, "ORDEN"))
        if not number:
            continue
        order_id = f"o-{number}"
        if order_id not in orders:
            operator_name = clean(value(row, "Aplicador")) or "Sin asignar"
            operator_id = slug(operator_name, "op")
            if operator_id not in operators:
                operators[operator_id] = {"id": operator_id, "name": operator_name, "phone": "", "active": True}
            orders[order_id] = {
                "id": order_id,
                "number": int(number) if isinstance(number, (int, float)) else number,
                "date": as_date(value(row, "FECHA")),
                "objective": clean(value(row, "Objetivo")),
                "crop": clean(value(row, "ESPECIE")),
                "variety": clean(value(row, "VARIEDAD")),
                "potrero": clean(value(row, "POTRERO")),
                "blocks": [part.strip() for part in str(clean(value(row, "BLOQUES"))).split(",") if part.strip()],
                "hectares": as_float(value(row, "HAS")),
                "waterHa": as_float(value(row, "Mojamiento/ HA Prog")),
                "speed": clean(value(row, "VELOCIDAD")),
                "nozzle": clean(value(row, "BOQUILLA")),
                "pressure": clean(value(row, "PRESION")),
                "tractorCode": clean(value(row, " CODtractor.")),
                "machineCode": clean(value(row, " COD maqui.")),
                "dosifier": clean(value(row, "Dossificador")),
                "operatorId": operator_id,
                "sprayerId": "eq-1",
                "tractorId": "tr-1",
                "status": "closed" if as_float(value(row, "LITROS APLICADOS")) else "planned",
                "notes": "Importado desde Aplicaciones.xlsx",
                "recipe": [],
                "tanks": [],
                "movements": [],
                "_importLiters": 0,
                "_importProducts": {},
            }

        product_name = clean(value(row, "Nombre  Comercial"))
        if product_name:
            product_id = ensure_product(
                state,
                product_ids,
                product_name,
                clean(value(row, "Ingrediente  Activo")),
                as_float(value(row, "Dosis (kg-lt/ 100Lts)")),
            )
            recipe = orders[order_id]["recipe"]
            if product_id and not any(item["productId"] == product_id for item in recipe):
                recipe.append({
                    "productId": product_id,
                    "dose100": as_float(value(row, "Dosis (kg-lt/ 100Lts)")),
                    "productHaProgram": as_float(value(row, "Gasto Producto / HA Prog")),
                    "totalProgram": as_float(value(row, "Total Producto HA (kG-LT)")),
                    "productHaReal": as_float(value(row, "GASTO PRODUCTO/HA REAL")),
                    "totalReal": as_float(value(row, "GASTO TOTAL PRODUCTO REAL")),
                })
            real_total = as_float(value(row, "GASTO TOTAL PRODUCTO REAL"))
            if real_total:
                orders[order_id]["_importProducts"][product_id] = real_total
            liters = as_float(value(row, "LITROS APLICADOS"))
            if liters:
                orders[order_id]["_importLiters"] = max(orders[order_id]["_importLiters"], liters)

    state["operators"] = list(operators.values()) or [{"id": "op-1", "name": "Sin asignar", "phone": "", "active": True}]
    for order in orders.values():
        if order["_importProducts"] or order["_importLiters"]:
            order["tanks"].append({
                "id": f"t-{order['number']}-excel",
                "liters": order["_importLiters"],
                "appliedAt": f"{order['date']}T08:00",
                "pressure": order["pressure"],
                "speed": order["speed"],
                "nozzle": order["nozzle"],
                "machineCode": order["machineCode"],
                "products": order["_importProducts"],
            })
        del order["_importLiters"]
        del order["_importProducts"]
    state["orders"] = list(orders.values())


def read_weather(wb, state):
    if "Estación Metereologica" not in wb.sheetnames:
        return
    ws = wb["Estación Metereologica"]
    headers = load_headers(ws, 2)
    by_day = defaultdict(lambda: {"max": [], "min": [], "wind": [], "humidity": []})

    def value(row, header):
        index = headers.get(header)
        return row[index] if index is not None and index < len(row) else None

    for row in ws.iter_rows(min_row=3, values_only=True):
        day = as_date(value(row, "Date"))
        by_day[day]["max"].append(as_float(value(row, "Hi temp")))
        by_day[day]["min"].append(as_float(value(row, "Low Temp")))
        by_day[day]["wind"].append(as_float(value(row, "Wind Speed")))
        by_day[day]["humidity"].append(as_float(value(row, "Humedad")) * 100)

    for day, data in list(by_day.items())[:400]:
        state["weather"].append({
            "date": day,
            "max": max(data["max"]) if data["max"] else 0,
            "min": min(data["min"]) if data["min"] else 0,
            "wind": round(sum(data["wind"]) / len(data["wind"]), 1) if data["wind"] else 0,
            "humidity": round(sum(data["humidity"]) / len(data["humidity"]), 1) if data["humidity"] else 0,
        })


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python tools/excel_to_seed.py Aplicaciones.xlsx respaldo.json")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    state = deepcopy(BASE)

    product_ids = read_products(wb, state)
    read_blocks(wb, state)
    read_orders(wb, state, product_ids)
    read_weather(wb, state)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Respaldo creado: {output_path}")
    print(f"Ordenes: {len(state['orders'])} | Productos: {len(state['products'])} | Bloques: {len(state['blocks'])}")


if __name__ == "__main__":
    main()
