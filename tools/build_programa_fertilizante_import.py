"""Genera el respaldo web y el SQL de importacion desde BASE DE DATOS BLOQUES."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\PC\Documents\FERTILIZANTE SUPA\Fertilizacion 2026-2027.xlsx")
SHEET = "BASE DE DATOS BLOQUES"
PRODUCT_ALIASES = {"HIBER HUMUS": "HIBER HUMUS 90PS"}
PRODUCT_UNITS = {
    "UREA": "KG",
    "FERPAC N22": "LT",
    "FERPAC FLUID 22": "LT",
    "SULFATO DE AMONIO": "KG",
    "CAPTURE 26": "KG",
    "MURIATO DE POTASIO": "KG",
    "ULTRASOL K ACID": "KG",
    "SULFATO DE MAGNESIO": "KG",
    "SULFATO DE ZINC": "KG",
    "ACIDO BORICO": "KG",
    "ULTRASOL PROP": "KG",
    "BIOAMINOL": "LT",
    "HIBER HUMUS 90PS": "KG",
    "MAGNIFIC FLOW CA": "LT",
    "B SOIL": "LT",
    "NITRATO DE CALCIO": "KG",
}
NUTRIENTS = ("n", "p", "k", "b", "zn", "mg", "ca", "ah", "af")


def clean_identifier(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def normalize_source_block(potrero: str, block: str) -> str:
    if potrero.upper() == "LOS PINOS PALTOS" and block.upper().startswith("A") and block[1:].isdigit():
        return block[1:]
    if potrero == "19" and block.upper() == "P19":
        return "4"
    return block


def build_records() -> list[dict]:
    workbook = load_workbook(SOURCE, data_only=True, read_only=True)
    worksheet = workbook[SHEET]
    compositions = {product: {nutrient: 0.0 for nutrient in NUTRIENTS} for product in PRODUCT_UNITS}
    for values in workbook["APORTES"].iter_rows(min_row=2, values_only=True):
        if not values[0]:
            continue
        product = PRODUCT_ALIASES.get(str(values[0]).strip(), str(values[0]).strip())
        if product in compositions:
            compositions[product] = {
                nutrient: float(values[index + 2] or 0)
                for index, nutrient in enumerate(NUTRIENTS)
            }
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    products = [str(value).strip() for value in headers[5:21]]
    grouped: dict[tuple[str, ...], dict] = defaultdict(lambda: {
        "cantidad_programada": 0.0,
        "hectareas_programadas": 0.0,
        "filas_excel": [],
    })
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        caseta, potrero, bloque, hectares, month = values[:5]
        if not caseta or not potrero or not bloque or not month:
            continue
        caseta = clean_identifier(caseta)
        potrero = clean_identifier(potrero)
        bloque = clean_identifier(bloque)
        bloque = normalize_source_block(potrero, bloque)
        month = str(month).strip()
        hectares = float(hectares or 0)
        for offset, source_product in enumerate(products):
            quantity = float(values[28 + offset] or 0)
            dose = float(values[5 + offset] or 0)
            if abs(quantity) < 1e-9 and abs(dose) < 1e-9:
                continue
            product = PRODUCT_ALIASES.get(source_product, source_product)
            key = (caseta, potrero, bloque, month, product)
            target = grouped[key]
            target["cantidad_programada"] += quantity
            target["filas_excel"].append(excel_row)
            target["hectareas_programadas"] += hectares

    records = []
    for (caseta, potrero, bloque, month, product), values in grouped.items():
        hectares = values["hectareas_programadas"]
        quantity = values["cantidad_programada"]
        record = {
            "caseta": caseta,
            "potrero": potrero,
            "bloque": bloque,
            "mes": month,
            "temporada": "2026-2027",
            "producto": product,
            "unidad": PRODUCT_UNITS[product],
            "dosis_por_ha": round(quantity / hectares, 6) if hectares else 0,
            "cantidad_programada": round(quantity, 6),
            "hectareas_programadas": round(hectares, 6),
            "filas_excel": ",".join(str(row) for row in values["filas_excel"]),
        }
        record.update(compositions.get(product, {}))
        records.append(record)
    return sorted(records, key=lambda row: (
        row["mes"], row["caseta"], row["potrero"], row["bloque"], row["producto"]
    ))


def sql_literal_json(records: list[dict]) -> str:
    return json.dumps(records, ensure_ascii=True, separators=(",", ":")).replace("'", "''")


def build_sql(records: list[dict]) -> str:
    source_json = sql_literal_json(records)
    products = """with source(nombre_comercial, nombre_normalizado, unidad, n, p, k, b, zn, mg, ca, ah, af, disolucion) as (values
  ('SULFATO DE AMONIO','SULFATO DE AMONIO','KG',0.21,0,0,0,0,0,0,0,0,0.20),
  ('CAPTURE 26','CAPTURE 26','KG',0.26,0,0,0,0,0,0,0,0,0.20),
  ('MAGNIFIC FLOW CA','MAGNIFIC FLOW CA','LT',0,0,0,0,0,0,0,0,0,0),
  ('B SOIL','B SOIL','LT',0.021,0.003,0.026,0,0,0.001,0.001,0.26,0.255,0.04),
  ('NITRATO DE CALCIO','NITRATO DE CALCIO','KG',0.16,0,0,0,0,0,0.165,0,0,0.15),
  ('HIBER HUMUS 90PS','HIBER HUMUS 90PS','KG',0,0,0.027,0,0,0,0,0.877,0,0.01)
)
insert into public.fertilizante_productos
  (nombre_comercial,nombre_normalizado,unidad,n,p,k,b,zn,mg,ca,ah,af,disolucion)
select * from source
on conflict (nombre_normalizado) do update set
  activo=true, actualizado_en=now();"""
    return f"""-- Generado desde {SOURCE.name}, hoja {SHEET}.
-- Ejecutar despues de supabase_programa_fertilizante.sql.
begin;

{products}

with source as (
  select * from jsonb_to_recordset('{source_json}'::jsonb) as x(
    caseta text, potrero text, bloque text, mes text, temporada text,
    producto text, unidad text, dosis_por_ha numeric,
    cantidad_programada numeric, hectareas_programadas numeric, filas_excel text
  )
), normalized as (
  select s.*,
    regexp_replace(regexp_replace(upper(trim(s.potrero)), '^P\\s*([0-9]+)$', '\\1'), '[^A-Z0-9]', '', 'g') as potrero_key,
    regexp_replace(regexp_replace(upper(trim(s.bloque)), '^(BLOQUE|B)\\s*', ''), '[^A-Z0-9]', '', 'g') as bloque_key,
    regexp_replace(upper(trim(s.caseta)), '[^A-Z0-9]', '', 'g') as caseta_key,
    regexp_replace(upper(trim(s.producto)), '[^A-Z0-9]', '', 'g') as producto_key
  from source s
), resolved as (
  select n.*, c.id as campo_id, fc.id as caseta_id, fp.id as producto_id
  from normalized n
  left join public.campos c on
    regexp_replace(regexp_replace(upper(trim(c.potrero)), '^P\\s*([0-9]+)$', '\\1'), '[^A-Z0-9]', '', 'g') = n.potrero_key
    and regexp_replace(regexp_replace(upper(trim(c.bloque)), '^(BLOQUE|B)\\s*', ''), '[^A-Z0-9]', '', 'g') = n.bloque_key
  left join public.fertilizante_casetas fc on
    regexp_replace(upper(trim(fc.nombre)), '[^A-Z0-9]', '', 'g') = n.caseta_key
  left join public.fertilizante_productos fp on
    regexp_replace(upper(trim(fp.nombre_normalizado)), '[^A-Z0-9]', '', 'g') = n.producto_key
), upserted as (
  insert into public.programa_fertilizante
    (campo_id,caseta_id,producto_id,mes,temporada,dosis_por_ha,cantidad_programada,
     hectareas_programadas,archivo_origen,filas_excel,actualizado_en)
  select campo_id,caseta_id,producto_id,(mes || '-01')::date,temporada,dosis_por_ha,
    cantidad_programada,hectareas_programadas,'{SOURCE.name}',filas_excel,now()
  from resolved
  where campo_id is not null and caseta_id is not null and producto_id is not null
  on conflict (campo_id,caseta_id,producto_id,mes) do update set
    temporada=excluded.temporada, dosis_por_ha=excluded.dosis_por_ha,
    cantidad_programada=excluded.cantidad_programada,
    hectareas_programadas=excluded.hectareas_programadas,
    archivo_origen=excluded.archivo_origen, filas_excel=excluded.filas_excel,
    actualizado_en=now()
  returning id
)
select
  (select count(*) from source) as filas_fuente,
  (select count(*) from upserted) as filas_insertadas_actualizadas,
  (select count(*) from resolved where campo_id is null) as sin_campo,
  (select count(*) from resolved where caseta_id is null) as sin_caseta,
  (select count(*) from resolved where producto_id is null) as sin_producto,
  coalesce((select jsonb_agg(distinct jsonb_build_object('potrero',potrero,'bloque',bloque)) from resolved where campo_id is null),'[]'::jsonb) as campos_sin_coincidencia,
  coalesce((select jsonb_agg(distinct caseta) from resolved where caseta_id is null),'[]'::jsonb) as casetas_sin_coincidencia,
  coalesce((select jsonb_agg(distinct producto) from resolved where producto_id is null),'[]'::jsonb) as productos_sin_coincidencia;

commit;
"""


def main() -> None:
    records = build_records()
    payload = {
        "source": SOURCE.name,
        "sheet": SHEET,
        "season": "2026-2027",
        "records": records,
    }
    (ROOT / "data").mkdir(exist_ok=True)
    (ROOT / "data" / "programa_fertilizante.json").write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    (ROOT / "supabase_programa_fertilizante_import.sql").write_text(
        build_sql(records), encoding="utf-8"
    )
    print(f"Generados {len(records)} registros consolidados")


if __name__ == "__main__":
    main()
