import re
import sys
from collections import defaultdict
from fechatime import fecha, fechatime
from pathlib import Path

import openpyxl


VALID_clasificacionS = {"N", "P", "VR", "ME", "VD", "M"}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def sql(value):
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (fechatime, fecha)):
        value = value.fecha() if isinstance(value, fechatime) else value
        return "'" + value.isoformat() + "'"
    return "'" + str(value).replace("'", "''") + "'"


def sql_text(value):
    if value is None or value == "":
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def sql_fecha(value):
    if isinstance(value, fechatime):
        return sql(value.fecha().isoformat())
    if isinstance(value, fecha):
        return sql(value.isoformat())
    if value:
        return sql(str(value)[:10])
    return "null"


def num(value, default=0):
    try:
        if value in (None, "", "nc", "NC"):
            return default
        if isinstance(value, str):
            match = re.search(r"-?\d+(?:[\.,]\d+)?", value)
            if match:
                return float(match.group(0).replace(",", "."))
        return float(value)
    except (tipoError, ValueError):
        return default


def integer(value):
    try:
        if value in (None, ""):
            return "null"
        return str(int(float(value)))
    except (tipoError, ValueError):
        return "null"


def headers(ws, row):
    return {clean(cell.value): idx for idx, cell in enumerate(ws[row]) if clean(cell.value)}


def get(row, idx, nombre, default=""):
    col = idx.get(nombre)
    if col is None or col >= len(row):
        return default
    return clean(row[col])


def season_years(nombre):
    match = re.search(r"(\d{4})\D+(\d{4})", str(nombre or ""))
    if match:
        return int(match.group(1)), int(match.group(2))
    return 2024, 2025


def normalize_clasificacion(value):
    text = str(value or "").strip().upper()
    if text in VALID_clasificacionS:
        return text
    if "NEB" in text:
        return "N"
    if "PUL" in text:
        return "P"
    if "RIEGO" in text:
        return "VR"
    if "ESPALDA" in text:
        return "ME"
    if "DRON" in text:
        return "VD"
    if "MANUAL" in text:
        return "M"
    return ""


def parse_bloques(value):
    if not value:
        return []
    return [part.strip() for part in str(value).replace(";", ",").split(",") if part.strip()]


def pg_array(values):
    return "array[" + ", ".join(sql(str(value)) for value in values) + "]::text[]"


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python tools/excel_to_supabase_sql.py Aplicaciones.xlsx carpeta_salida_o_salida.sql")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)

    temporadas = {}
    productos = {}
    campos = {}
    programas = {}
    orders = {}
    order_productos = {}

    if "Productos" in wb.sheetnombres:
        ws = wb["Productos"]
        h = headers(ws, 2)
        for row in ws.iter_rows(min_row=3, values_only=True):
            nombre = get(row, h, "NOMBRE COMERCIAL")
            if not nombre:
                continue
            key = nombre.lower()
            productos[key] = {
                "nombre": nombre,
                "ingrediente_activo": get(row, h, "ingrediente_activoE ACTIVO"),
                "unidad": "kg",
                "dosis_por_100": num(get(row, h, "Dosis /100lts    (grs-cc)")),
                "horas_reingreso": 24,
                "dias_carencia": num(get(row, h, "PerÃ­odo Carencia Agenda de Pesticidas")),
            }

    if "HAS Canelillo" in wb.sheetnombres:
        ws = wb["HAS Canelillo"]
        h = headers(ws, 2)
        for row in ws.iter_rows(min_row=3, values_only=True):
            potrero = get(row, h, "Potrero")
            block = get(row, h, "Bloque")
            if not potrero or block == "":
                continue
            campos[(str(potrero), str(block))] = {
                "potrero": str(potrero),
                "block": str(block),
                "cultivo": get(row, h, "Especie"),
                "variedad": get(row, h, "Variedad"),
                "hectareas": num(get(row, h, "Hectareas")),
            }

    ws = wb["Aplicaciones"]
    h = headers(ws, 5)
    for row in ws.iter_rows(min_row=6, values_only=True):
        numero_orden = get(row, h, "ORDEN")
        if numero_orden == "":
            continue
        season_nombre = get(row, h, "TEMPORADA") or "2024/2025"
        anio_inicio, anio_fin = season_years(season_nombre)
        temporadas[season_nombre] = {"nombre": season_nombre, "anio_inicio": anio_inicio, "anio_fin": anio_fin}
        numero_programa_raw = get(row, h, "NÂ°Prog")
        numero_programa = int(num(numero_programa_raw)) if numero_programa_raw not in ("", None) else None
        objetivo = get(row, h, "Objetivo")
        if numero_programa:
            programas[(season_nombre, numero_programa)] = {
                "season": season_nombre,
                "number": numero_programa,
                "nombre": objetivo or f"Programa {numero_programa}",
                "cultivo": get(row, h, "ESPECIE"),
                "objetivo": objetivo,
            }

        numero_orden_value = num(numero_orden)
        key = (season_nombre, numero_orden_value)
        if key not in orders:
            orders[key] = {
                "season": season_nombre,
                "numero_orden": numero_orden_value,
                "numero_programa": numero_programa,
                "nombre_programa": objetivo,
                "clasificacion": normalize_clasificacion(get(row, h, "CLASIFICACION")),
                "fecha": get(row, h, "FECHA"),
                "fecha_planificada": get(row, h, "FECHA") or fecha.today().isoformat(),
                "objetivo": objetivo,
                "cultivo": get(row, h, "ESPECIE"),
                "variedad": get(row, h, "VARIEDAD"),
                "potrero": get(row, h, "POTRERO"),
                "bloques": parse_bloques(get(row, h, "BLOQUES")),
                "hectareas": num(get(row, h, "HAS")),
                "agua_por_ha": num(get(row, h, "Mojamiento/ HA Prog")),
                "presion": num(get(row, h, "PRESION"), None),
                "boquilla": get(row, h, "BOQUILLA"),
                "velocidad": num(get(row, h, "VELOCIDAD"), None),
                "codigo_tractor": get(row, h, " CODtractor."),
                "codigo_maquina": get(row, h, " COD maqui."),
                "dosificador": get(row, h, "Dossificador"),
                "estado": "completada" if num(get(row, h, "LITROS APLICADOS")) else "planificada",
                "dispatch_litros": 0,
                "despacho_productos": {},
            }

        product_nombre = get(row, h, "Nombre  Comercial")
        if product_nombre:
            product_key = product_nombre.lower()
            productos.setdefault(product_key, {
                "nombre": product_nombre,
                "ingrediente_activo": get(row, h, "ingrediente_activoe  Activo"),
                "unidad": "kg",
                "dosis_por_100": num(get(row, h, "Dosis (kg-lt/ 100Lts)")),
                "horas_reingreso": 24,
                "dias_carencia": 0,
            })
            order_productos[(key, product_key)] = {
                "order": key,
                "product": product_key,
                "dosis_por_100": num(get(row, h, "Dosis (kg-lt/ 100Lts)")),
                "producto_por_ha_programa": num(get(row, h, "Gasto Producto / HA Prog")),
                "total_programa": num(get(row, h, "Total Producto HA (kG-LT)")),
            }
            real_total = num(get(row, h, "GASTO TOTAL PRODUCTO REAL"), 0)
            if real_total:
                orders[key]["despacho_productos"][product_key] = real_total
        litros = num(get(row, h, "LITROS APLICADOS"), 0)
        if litros:
            orders[key]["dispatch_litros"] = max(orders[key]["dispatch_litros"], litros)

    pre_lines = [
        "-- 00_preparar_import.sql",
        "-- Ejecutar primero.",
        "alter table public.ordenes_aplicacion add column if not exists clasificacion text;",
        "alter table public.ordenes_aplicacion alter column numero_orden type numeric(12,2) using numero_orden::numeric;",
        "alter table public.ordenes_aplicacion drop constraint if exists ordenes_aplicacion_clasificacion_check;",
        "alter table public.ordenes_aplicacion add constraint ordenes_aplicacion_clasificacion_check check (clasificacion is null or clasificacion in ('N','P','VR','ME','VD','M'));",
        "delete from public.movimientos_stock where nota like 'Importado Excel%';",
        "delete from public.despachos where nota like 'Importado Excel%';",
    ]

    catalog_lines = ["-- 01_catalogos.sql", "-- Temporadas, productos, potreros y programas."]
    order_lines = ["-- 02_ordenes.sql", "-- Ordenes de aplicacion."]
    recipe_lines = ["-- 03_recetas.sql", "-- Productos por orden."]
    dispatch_lines = ["-- 04_salidas_historicas.sql", "-- Salidas historicas importadas desde Excel."]

    for season in temporadas.values():
        catalog_lines.append(
            f"insert into public.temporadas (nombre,anio_inicio,anio_fin,estado) values ({sql(season['nombre'])},{season['anio_inicio']},{season['anio_fin']},'activa') "
            "on conflict (nombre) do update set anio_inicio=excluded.anio_inicio,anio_fin=excluded.anio_fin,estado=excluded.estado;"
        )

    for product in productos.values():
        catalog_lines.append(
            "insert into public.productos (nombre,ingrediente_activo,unidad,dosis_por_100,horas_reingreso,dias_carencia) values "
            f"({sql_text(product['nombre'])},{sql_text(product['ingrediente_activo'])},{sql_text(product['unidad'])},{product['dosis_por_100']},{int(product['horas_reingreso'])},{int(product['dias_carencia'])}) "
            "on conflict (nombre) do update set ingrediente_activo=excluded.ingrediente_activo, unidad=excluded.unidad, dosis_por_100=excluded.dosis_por_100, horas_reingreso=excluded.horas_reingreso, dias_carencia=excluded.dias_carencia;"
        )

    for field in campos.values():
        catalog_lines.append(
            "insert into public.campos (potrero,bloque,cultivo,variedad,hectareas,activo) values "
            f"({sql_text(field['potrero'])},{sql_text(field['block'])},{sql_text(field['cultivo'])},{sql_text(field['variedad'])},{field['hectareas']},true) "
            "on conflict (potrero,bloque) do update set cultivo=excluded.cultivo,variedad=excluded.variedad,hectareas=excluded.hectareas,activo=true;"
        )

    for program in programas.values():
        catalog_lines.append(
            "insert into public.programas (temporada_id,numero_programa,nombre,cultivo,objetivo) "
            f"select s.id,{program['number']},{sql_text(program['nombre'])},{sql_text(program['cultivo'])},{sql_text(program['objetivo'])} from public.temporadas s where s.nombre={sql_text(program['season'])} "
            "on conflict (temporada_id,numero_programa) do update set nombre=excluded.nombre,cultivo=excluded.cultivo,objetivo=excluded.objetivo;"
        )

    for order in orders.values():
        program_join = f"left join public.programas p on p.temporada_id=s.id and p.numero_programa={order['numero_programa']}" if order["numero_programa"] else "left join public.programas p on false"
        order_lines.append(
            "insert into public.ordenes_aplicacion (temporada_id,programa_id,numero_orden,numero_programa,nombre_programa,clasificacion,fecha,fecha_planificada,objetivo,cultivo,variedad,potrero,bloques,hectareas,agua_por_ha,presion,boquilla,velocidad,codigo_tractor,codigo_maquina,dosificador,estado) "
            f"select s.id,p.id,{order['numero_orden']},{order['numero_programa'] if order['numero_programa'] else 'null'},{sql_text(order['nombre_programa'])},{sql_text(order['clasificacion'])},{sql_fecha(order['fecha'])},{sql_fecha(order['fecha_planificada'])},{sql_text(order['objetivo'])},{sql_text(order['cultivo'])},{sql_text(order['variedad'])},{sql_text(order['potrero'])},{pg_array(order['bloques'])},{order['hectareas']},{order['agua_por_ha']},{sql(order['presion'])},{sql_text(order['boquilla'])},{sql(order['velocidad'])},{sql_text(order['codigo_tractor'])},{sql_text(order['codigo_maquina'])},{sql_text(order['dosificador'])},{sql_text(order['estado'])} "
            f"from public.temporadas s {program_join} where s.nombre={sql_text(order['season'])} "
            "on conflict (temporada_id,numero_orden) do update set programa_id=excluded.programa_id,numero_programa=excluded.numero_programa,nombre_programa=excluded.nombre_programa,clasificacion=excluded.clasificacion,fecha=excluded.fecha,fecha_planificada=excluded.fecha_planificada,objetivo=excluded.objetivo,cultivo=excluded.cultivo,variedad=excluded.variedad,potrero=excluded.potrero,bloques=excluded.bloques,hectareas=excluded.hectareas,agua_por_ha=excluded.agua_por_ha,presion=excluded.presion,boquilla=excluded.boquilla,velocidad=excluded.velocidad,codigo_tractor=excluded.codigo_tractor,codigo_maquina=excluded.codigo_maquina,dosificador=excluded.dosificador,estado=excluded.estado;"
        )

    for row in order_productos.values():
        season_nombre, numero_orden = row["order"]
        product_nombre = productos[row["product"]]["nombre"]
        recipe_lines.append(
            "insert into public.orden_productos (orden_id,producto_id,numero_programa,dosis_por_100,producto_por_ha_programa,total_programa) "
            f"select o.id,p.id,o.numero_programa,{row['dosis_por_100']},{row['producto_por_ha_programa']},{row['total_programa']} "
            "from public.ordenes_aplicacion o "
            "join public.temporadas s on s.id=o.temporada_id "
            f"join public.productos p on p.nombre={sql_text(product_nombre)} "
            f"where s.nombre={sql_text(season_nombre)} and o.numero_orden={numero_orden} "
            "on conflict (orden_id,producto_id,numero_programa) do update set dosis_por_100=excluded.dosis_por_100,producto_por_ha_programa=excluded.producto_por_ha_programa,total_programa=excluded.total_programa;"
        )

    for order in orders.values():
        if not order["dispatch_litros"] and not order["despacho_productos"]:
            continue
        values = []
        for product_key, cantidad in order["despacho_productos"].items():
            product_nombre = productos[product_key]["nombre"]
            values.append(f"({sql_text(product_nombre)},{cantidad})")
        if not values:
            continue
        program_join = ""
        dispatch_lines.append(
            "with ord as ("
            "select o.id as orden_id from public.ordenes_aplicacion o join public.temporadas s on s.id=o.temporada_id "
            f"where s.nombre={sql_text(order['season'])} and o.numero_orden={order['numero_orden']} limit 1"
            "), ins as ("
            "insert into public.despachos (orden_id,tipo,fecha,litros,nota) "
            f"select orden_id,'salida',{sql_fecha(order['fecha'])},{order['dispatch_litros']},{sql_text('Importado Excel orden ' + str(order['numero_orden']))} from ord returning id,orden_id"
            "), vals(product_nombre,cantidad) as (values " + ",".join(values) + "), dp as ("
            "insert into public.despacho_productos (despacho_id,producto_id,cantidad,costo_unitario,lote) "
            "select ins.id,p.id,vals.cantidad,p.costo_unitario,p.lote from ins cross join vals join public.productos p on p.nombre=vals.product_nombre "
            "returning despacho_id,producto_id,cantidad,costo_unitario,lote"
            ") "
            "insert into public.movimientos_stock (producto_id,orden_id,despacho_id,tipo,fecha,cantidad,costo_unitario,lote,nota) "
            f"select dp.producto_id,ins.orden_id,dp.despacho_id,'salida',{sql_fecha(order['fecha'])},dp.cantidad,dp.costo_unitario,dp.lote,{sql_text('Importado Excel orden ' + str(order['numero_orden']))} "
            "from dp join ins on ins.id=dp.despacho_id;"
        )

    verify_lines = [
        "-- 05_verificar_import.sql",
        "select 'temporadas' as tabla, count(*) from public.temporadas;",
        "select 'productos' as tabla, count(*) from public.productos;",
        "select 'campos' as tabla, count(*) from public.campos;",
        "select 'programas' as tabla, count(*) from public.programas;",
        "select 'ordenes_aplicacion' as tabla, count(*) from public.ordenes_aplicacion;",
        "select 'orden_productos' as tabla, count(*) from public.orden_productos;",
        "select 'despachos' as tabla, count(*) from public.despachos;",
        "select 'despacho_productos' as tabla, count(*) from public.despacho_productos;",
        "select 'movimientos_stock' as tabla, count(*) from public.movimientos_stock;",
    ]

    if output_path.suffix.lower() == ".sql":
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("\n\n".join([
            "\n".join(pre_lines),
            "\n".join(catalog_lines),
            "\n".join(order_lines),
            "\n".join(recipe_lines),
            "\n".join(dispatch_lines),
            "\n".join(verify_lines),
        ]), encoding="utf-8")
        print(f"SQL generado: {output_path}")
    else:
        output_path.mkdir(parents=True, exist_ok=True)
        files = {
            "00_preparar_import.sql": pre_lines,
            "01_catalogos.sql": catalog_lines,
            "02_ordenes.sql": order_lines,
            "03_recetas.sql": recipe_lines,
            "04_salidas_historicas.sql": dispatch_lines,
            "05_verificar_import.sql": verify_lines,
        }
        for nombre, content in files.items():
            (output_path / nombre).write_text("\n".join(content), encoding="utf-8")
        print(f"SQL generado en carpeta: {output_path}")
    print(f"Temporadas: {len(temporadas)} | Productos: {len(productos)} | Potreros/bloques: {len(campos)} | Ordenes: {len(orders)} | Recetas: {len(order_productos)}")


if __nombre__ == "__main__":
    main()


