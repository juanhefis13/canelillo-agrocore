from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_NAME = "COSECHA SUPA.xlsx"
BATCH_SIZE = 500


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value) -> str:
    text = clean_text(value).replace("\xa0", " ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_label(value) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip().upper()


def normalize_potrero(value) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def parse_number(value, default=0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    is_percent = "%" in text
    text = text.replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    number = float(match.group(0))
    return number / 100 if is_percent else number


def parse_int(value):
    number = parse_number(value, None)
    if number is None:
        return None
    return int(number)


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def sql_text(value, default_null=True) -> str:
    text = clean_text(value)
    if not text and default_null:
        return "null"
    return "'" + text.replace("'", "''") + "'"


def sql_date(value) -> str:
    parsed = parse_date(value)
    return "null" if not parsed else f"'{parsed.isoformat()}'"


def format_number(value, digits=6) -> str:
    number = parse_number(value, 0)
    if abs(number) < 0.000000001:
        return "0"
    text = f"{number:.{digits}f}".rstrip("0").rstrip(".")
    return text or "0"


def sql_num(value, default=0, digits=6) -> str:
    if value is None or value == "":
        return "null" if default is None else format_number(default, digits)
    return format_number(value, digits)


def sql_jsonb(values: dict[str, float]) -> str:
    if not values:
        return "'{}'::jsonb"
    payload = {key: round(value, 6) for key, value in values.items() if abs(value) >= 0.000000001}
    text = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return "'" + text.replace("'", "''") + "'::jsonb"


def headers_for(row) -> dict[str, int]:
    return {normalize_header(value): index for index, value in enumerate(row) if clean_text(value)}


def find_index(header_index: dict[str, int], *aliases: str) -> int:
    for alias in aliases:
        key = normalize_header(alias)
        if key in header_index:
            return header_index[key]
    raise KeyError(f"No se encontro columna: {aliases}")


def value(row, index):
    return row[index] if index < len(row) else None


def export_percent(value):
    number = parse_number(value, None)
    if number is None:
        return None
    return number / 100 if abs(number) > 1 else number


def calculated_export_percent(exported_kg, kg_en_proceso):
    denominator = parse_number(kg_en_proceso, 0)
    if not denominator:
        return None
    return parse_number(exported_kg, 0) / denominator


def calibre_label(header, is_caja=False) -> str:
    text = re.sub(r"\s+", " ", clean_text(header)).strip().upper()
    if is_caja:
        text = re.sub(r"^(CAJAS?|CAJ)\s+", "", text)
        text = re.sub(r"^CAT\s+(\d+)", r"\1", text)
        text = re.sub(r"^C(\d+)$", r"\1", text)
    return text


def detect_species_by_variety(cosecha_rows):
    counts: dict[str, Counter] = defaultdict(Counter)
    for row in cosecha_rows:
        species = normalize_label(row["especie"])
        variety = normalize_label(row["variedad"])
        if species and variety:
            counts[variety][species] += 1
    mapping = {variety: counter.most_common(1)[0][0] for variety, counter in counts.items()}
    mapping.setdefault("BARNFIELD", "NARANJA")
    return mapping


def read_cosecha_sheet(wb):
    ws = wb["BD COSECHA SUPA"]
    iterator = ws.iter_rows(values_only=True)
    header_index = headers_for(next(iterator))
    idx = {
        "especie": find_index(header_index, "Especie"),
        "variedad": find_index(header_index, "Variedad"),
        "potrero": find_index(header_index, "Potrero"),
        "bloque_formula": find_index(header_index, "Bloque Formula"),
        "bloque": find_index(header_index, "Bloque"),
        "fecha": find_index(header_index, "FECHA COSECHA"),
        "semana": find_index(header_index, "Semana"),
        "contratista": find_index(header_index, "CONTRATISTA"),
        "cuadrilla": find_index(header_index, "CUADRILLA"),
        "jornales": find_index(header_index, "JORNALES"),
        "bins_nac": find_index(header_index, "BINS NAC"),
        "bins_expo": find_index(header_index, "BINS EXPO"),
        "total_bins": find_index(header_index, "Total Bins"),
        "kg_nac": find_index(header_index, "KG NAC"),
        "kg_exp": find_index(header_index, "KG EXP"),
        "kg_totales": find_index(header_index, "KG TOTALES"),
    }
    rows = []
    for fila_excel, row in enumerate(iterator, start=2):
        fecha = parse_date(value(row, idx["fecha"]))
        especie = normalize_label(value(row, idx["especie"]))
        variedad = normalize_label(value(row, idx["variedad"]))
        potrero = normalize_potrero(value(row, idx["potrero"]))
        if not fecha or not variedad or not potrero:
            continue
        bloque_formula = normalize_potrero(value(row, idx["bloque_formula"]))
        bloque_excel = normalize_potrero(value(row, idx["bloque"]))
        bloque_normalizado = bloque_excel or bloque_formula
        rows.append({
            "fecha": fecha,
            "anio": fecha.year,
            "semana": parse_int(value(row, idx["semana"])),
            "especie": especie,
            "variedad": variedad,
            "potrero_excel": potrero,
            "bloque_formula": bloque_formula,
            "bloque_excel": bloque_excel,
            "potrero_normalizado": potrero,
            "bloque_normalizado": bloque_normalizado,
            "contratista": normalize_potrero(value(row, idx["contratista"])),
            "cuadrilla": normalize_potrero(value(row, idx["cuadrilla"])),
            "jornales": parse_number(value(row, idx["jornales"]), 0),
            "bins_nac": parse_number(value(row, idx["bins_nac"]), 0),
            "bins_expo": parse_number(value(row, idx["bins_expo"]), 0),
            "total_bins": parse_number(value(row, idx["total_bins"]), 0),
            "kg_nac": parse_number(value(row, idx["kg_nac"]), 0),
            "kg_exp": parse_number(value(row, idx["kg_exp"]), 0),
            "kg_totales": parse_number(value(row, idx["kg_totales"]), 0),
            "archivo_origen": SOURCE_NAME,
            "fila_excel": fila_excel,
        })
    return rows


def read_export_sheet(wb, species_by_variety):
    ws = wb["BD EXPORTACION SUPA"]
    iterator = ws.iter_rows(values_only=True)
    header_row = list(next(iterator))
    header_index = headers_for(header_row)
    idx = {
        "variedad": find_index(header_index, "VARIEDAD"),
        "anio": find_index(header_index, "ANO", "AÑO"),
        "fecha": find_index(header_index, "Fecha"),
        "potrero": find_index(header_index, "Potrero"),
        "cant_bins": find_index(header_index, "Cant Bins"),
        "enviados_kg": find_index(header_index, "Enviados Kg"),
        "recepcionados_kg": find_index(header_index, "Recepcionados"),
        "diferencia_kg": find_index(header_index, "dif"),
        "bins_por_confirmar": find_index(header_index, "Bins por confirmar"),
        "kg_en_proceso": find_index(header_index, "Kg. I Proceso"),
        "kg_por_procesar": find_index(header_index, "Kg. X Procesar", "Kg. X  Procesar"),
        "exportados_kg": find_index(header_index, "exportados"),
        "descarte_kg": find_index(header_index, "descarte"),
        "precalibre_kg": find_index(header_index, "precalibre"),
        "desecho_kg": find_index(header_index, "Desecho"),
        "merma_kg": find_index(header_index, "merma"),
        "x_kg": find_index(header_index, "x."),
        "porcentaje_expo": find_index(header_index, "% Expo"),
        "cajas": find_index(header_index, "cajas"),
    }
    kg_cols = list(range(idx["porcentaje_expo"] + 1, idx["cajas"]))
    caja_cols = list(range(idx["cajas"] + 1, len(header_row)))
    rows = []
    for fila_excel, row in enumerate(iterator, start=2):
        variedad = normalize_label(value(row, idx["variedad"]))
        fecha = parse_date(value(row, idx["fecha"]))
        potrero = normalize_potrero(value(row, idx["potrero"]))
        if not variedad or not fecha or not potrero:
            continue
        anio = parse_int(value(row, idx["anio"])) or fecha.year
        calibres_kg = {}
        calibres_cajas = {}
        for col in kg_cols:
            amount = parse_number(value(row, col), 0)
            label = calibre_label(header_row[col], is_caja=False)
            if label and amount:
                calibres_kg[label] = calibres_kg.get(label, 0) + amount
        for col in caja_cols:
            amount = parse_number(value(row, col), 0)
            label = calibre_label(header_row[col], is_caja=True)
            if label and amount:
                calibres_cajas[label] = calibres_cajas.get(label, 0) + amount
        kg_en_proceso = parse_number(value(row, idx["kg_en_proceso"]), 0)
        exportados_kg = parse_number(value(row, idx["exportados_kg"]), 0)
        rows.append({
            "fecha": fecha,
            "anio": anio,
            "especie": species_by_variety.get(variedad, "SIN ESPECIE"),
            "variedad": variedad,
            "potrero_excel": potrero,
            "potrero_normalizado": potrero,
            "cant_bins": parse_number(value(row, idx["cant_bins"]), 0),
            "enviados_kg": parse_number(value(row, idx["enviados_kg"]), 0),
            "recepcionados_kg": parse_number(value(row, idx["recepcionados_kg"]), 0),
            "diferencia_kg": parse_number(value(row, idx["diferencia_kg"]), 0),
            "bins_por_confirmar": parse_number(value(row, idx["bins_por_confirmar"]), 0),
            "kg_en_proceso": kg_en_proceso,
            "kg_por_procesar": parse_number(value(row, idx["kg_por_procesar"]), 0),
            "exportados_kg": exportados_kg,
            "descarte_kg": parse_number(value(row, idx["descarte_kg"]), 0),
            "precalibre_kg": parse_number(value(row, idx["precalibre_kg"]), 0),
            "desecho_kg": parse_number(value(row, idx["desecho_kg"]), 0),
            "merma_kg": parse_number(value(row, idx["merma_kg"]), 0),
            "x_kg": parse_number(value(row, idx["x_kg"]), 0),
            "porcentaje_expo": calculated_export_percent(exportados_kg, kg_en_proceso),
            "calibres_kg": calibres_kg,
            "calibres_cajas": calibres_cajas,
            "calibres_kg_total": sum(calibres_kg.values()),
            "calibres_cajas_total": sum(calibres_cajas.values()),
            "archivo_origen": SOURCE_NAME,
            "fila_excel": fila_excel,
        })
    return rows


def row_sql(row, columns):
    values = []
    for column in columns:
        raw = row[column]
        if column in {"fecha"}:
            values.append(sql_date(raw))
        elif column in {"anio", "semana", "fila_excel"}:
            values.append("null" if raw is None else str(int(raw)))
        elif column in {"calibres_kg", "calibres_cajas"}:
            values.append(sql_jsonb(raw))
        elif column == "porcentaje_expo":
            values.append(sql_num(raw, None, 6))
        elif column in {
            "jornales", "bins_nac", "bins_expo", "total_bins", "kg_nac", "kg_exp", "kg_totales",
            "cant_bins", "enviados_kg", "recepcionados_kg", "diferencia_kg", "bins_por_confirmar",
            "kg_en_proceso", "kg_por_procesar", "exportados_kg", "descarte_kg", "precalibre_kg",
            "desecho_kg", "merma_kg", "x_kg", "calibres_kg_total", "calibres_cajas_total",
        }:
            values.append(sql_num(raw, 0, 6))
        else:
            values.append(sql_text(raw, default_null=True))
    return "(" + ", ".join(values) + ")"


def write_batches(handle, table, columns, rows, extra_updates=None):
    if not rows:
        return
    update_columns = [column for column in columns if column not in {"archivo_origen", "fila_excel"}]
    updates = [f"{column} = excluded.{column}" for column in update_columns]
    updates.extend(extra_updates or [])
    for start in range(0, len(rows), BATCH_SIZE):
        batch = rows[start:start + BATCH_SIZE]
        handle.write(f"insert into public.{table} ({', '.join(columns)}) values\n")
        handle.write(",\n".join(row_sql(row, columns) for row in batch))
        handle.write("\non conflict (archivo_origen, fila_excel) do update set\n  ")
        handle.write(",\n  ".join(updates))
        handle.write(";\n\n")


def write_normalization(handle):
    handle.write(
        """-- Normaliza cosecha/exportacion contra public.campos.
create or replace function public.agrocore_normalizar_campo_ref(valor text, tipo text default 'general')
returns text
language sql
immutable
as $$
  with base as (
    select trim(
      regexp_replace(
        regexp_replace(
          regexp_replace(lower(coalesce(valor, '')), '\\s+', ' ', 'g'),
          '^(potrero|bloque|cuartel)\\s+',
          '',
          'g'
        ),
        '\\s+',
        ' ',
        'g'
      )
    ) as texto
  )
  select nullif(
    case
      when tipo = 'potrero' then regexp_replace(texto, '^p\\s*([0-9])', '\\1', 'g')
      when tipo = 'bloque' then regexp_replace(texto, '^b\\s*([0-9])', '\\1', 'g')
      else texto
    end,
    ''
  )
  from base;
$$;

update public.cosecha_analisis ca
set campo_id = c.id,
    potrero_normalizado = c.potrero,
    bloque_normalizado = c.bloque
from public.campos c
where public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') = public.agrocore_normalizar_campo_ref(coalesce(ca.potrero_normalizado, ca.potrero_excel), 'potrero')
  and public.agrocore_normalizar_campo_ref(c.bloque, 'bloque') = public.agrocore_normalizar_campo_ref(coalesce(ca.bloque_excel, ca.bloque_normalizado, ca.bloque_formula), 'bloque');

with matches as (
  select ca.id as cosecha_id, c.id as campo_id, c.potrero, c.bloque
  from public.cosecha_analisis ca
  join public.campos c
    on public.agrocore_normalizar_campo_ref(coalesce(ca.potrero_normalizado, ca.potrero_excel), 'potrero') = '26'
   and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') = lower(substring(public.agrocore_normalizar_campo_ref(coalesce(ca.bloque_excel, ca.bloque_normalizado, ca.bloque_formula), 'bloque') from '^([a-z])\\s*0*[0-9]+$'))
   and (
        public.agrocore_normalizar_campo_ref(c.bloque, 'bloque') = public.agrocore_normalizar_campo_ref(coalesce(ca.bloque_excel, ca.bloque_normalizado, ca.bloque_formula), 'bloque')
        or public.agrocore_normalizar_campo_ref(c.bloque, 'bloque') = regexp_replace(public.agrocore_normalizar_campo_ref(coalesce(ca.bloque_excel, ca.bloque_normalizado, ca.bloque_formula), 'bloque'), '^[a-z]\\s*0*([0-9]+)$', '\\1')
      )
)
update public.cosecha_analisis ca
set campo_id = m.campo_id,
    potrero_normalizado = m.potrero,
    bloque_normalizado = m.bloque
from matches m
where ca.id = m.cosecha_id;

with matches as (
  select
    ca.id as cosecha_id,
    (array_agg(c.id order by c.id::text))[1] as campo_id,
    (array_agg(c.potrero order by c.id::text))[1] as potrero,
    (array_agg(c.bloque order by c.id::text))[1] as bloque,
    count(*) as coincidencias
  from public.cosecha_analisis ca
  join public.campos c
    on (
        public.agrocore_normalizar_campo_ref(ca.potrero_excel, 'potrero') = '27'
        or public.agrocore_normalizar_campo_ref(ca.potrero_normalizado, 'potrero') = '27'
      )
   and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') like '27%'
   and public.agrocore_normalizar_campo_ref(c.bloque, 'bloque') = public.agrocore_normalizar_campo_ref(coalesce(ca.bloque_excel, ca.bloque_normalizado, ca.bloque_formula), 'bloque')
   and public.agrocore_normalizar_campo_ref(c.variedad) = public.agrocore_normalizar_campo_ref(ca.variedad)
   and (
        (
          public.agrocore_normalizar_campo_ref(ca.bloque_formula, 'bloque') like 'oo%'
          and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') = '27 imp'
        )
        or (
          public.agrocore_normalizar_campo_ref(ca.bloque_formula, 'bloque') like 'o%'
          and public.agrocore_normalizar_campo_ref(ca.bloque_formula, 'bloque') not like 'oo%'
          and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') = '27 grav'
        )
      )
  group by ca.id
  having count(*) = 1
)
update public.cosecha_analisis ca
set campo_id = m.campo_id,
    potrero_normalizado = m.potrero,
    bloque_normalizado = m.bloque
from matches m
where ca.id = m.cosecha_id;

with matches as (
  select
    ea.id as exportacion_id,
    array_agg(c.id order by c.potrero, c.bloque) as campo_ids,
    string_agg(distinct c.especie, ' / ' order by c.especie) as especie_match
  from public.exportacion_analisis ea
  join public.campos c
    on public.agrocore_normalizar_campo_ref(c.variedad) = public.agrocore_normalizar_campo_ref(ea.variedad)
   and (
        public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') = public.agrocore_normalizar_campo_ref(coalesce(ea.potrero_normalizado, ea.potrero_excel), 'potrero')
        or (public.agrocore_normalizar_campo_ref(coalesce(ea.potrero_normalizado, ea.potrero_excel), 'potrero') = '27' and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') like '27%')
        or (public.agrocore_normalizar_campo_ref(coalesce(ea.potrero_normalizado, ea.potrero_excel), 'potrero') = '26' and public.agrocore_normalizar_campo_ref(c.potrero, 'potrero') in ('d','e','f','g','h','i','j'))
      )
  group by ea.id
)
update public.exportacion_analisis ea
set campo_ids = m.campo_ids,
    especie = coalesce(nullif(ea.especie, ''), m.especie_match)
from matches m
where ea.id = m.exportacion_id;
"""
    )


def main():
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        raise SystemExit("Uso: python tools/build_cosecha_analisis_sql.py COSECHA_SUPA.xlsx [salida.sql]")
    global SOURCE_NAME
    input_path = Path(sys.argv[1])
    SOURCE_NAME = input_path.name
    output_path = Path(sys.argv[2]) if len(sys.argv) == 3 else Path("supabase_cosecha_analisis_import.sql")
    wb = load_workbook(input_path, read_only=True, data_only=True)
    cosecha_rows = read_cosecha_sheet(wb)
    species_by_variety = detect_species_by_variety(cosecha_rows)
    export_rows = read_export_sheet(wb, species_by_variety)

    cosecha_columns = [
        "fecha", "anio", "semana", "especie", "variedad", "potrero_excel", "bloque_formula",
        "bloque_excel", "potrero_normalizado", "bloque_normalizado", "contratista", "cuadrilla",
        "jornales", "bins_nac", "bins_expo", "total_bins", "kg_nac", "kg_exp", "kg_totales",
        "archivo_origen", "fila_excel",
    ]
    export_columns = [
        "fecha", "anio", "especie", "variedad", "potrero_excel", "potrero_normalizado",
        "cant_bins", "enviados_kg", "recepcionados_kg", "diferencia_kg", "bins_por_confirmar",
        "kg_en_proceso", "kg_por_procesar", "exportados_kg", "descarte_kg", "precalibre_kg",
        "desecho_kg", "merma_kg", "x_kg", "porcentaje_expo", "calibres_kg", "calibres_cajas",
        "calibres_kg_total", "calibres_cajas_total", "archivo_origen", "fila_excel",
    ]

    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("-- Import generado desde COSECHA SUPA.xlsx.\n")
        handle.write("-- Ejecutar despues de supabase_cosecha_analisis.sql.\n")
        handle.write("-- Corrige/actualiza registros por archivo_origen + fila_excel sin cambiar sus ids.\n\n")
        handle.write("begin;\n\n")
        write_batches(handle, "cosecha_analisis", cosecha_columns, cosecha_rows, extra_updates=["campo_id = null"])
        write_batches(handle, "exportacion_analisis", export_columns, export_rows, extra_updates=["campo_ids = null"])
        write_normalization(handle)
        handle.write("\ncommit;\n")

    desecho_total = sum(row["desecho_kg"] for row in export_rows)
    print(f"cosecha_rows={len(cosecha_rows)}")
    print(f"export_rows={len(export_rows)}")
    print(f"desecho_total={desecho_total:.3f}")
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
